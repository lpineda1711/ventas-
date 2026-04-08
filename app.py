import streamlit as st
import pandas as pd
import pdfplumber
import re
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side

st.title("📊 Ventas SRI - Formato PRO")

uploaded_files = st.file_uploader(
    "Sube facturas PDF",
    type="pdf",
    accept_multiple_files=True
)

def limpiar_numero(texto):
    if not texto:
        return 0
    texto = texto.replace(",", ".")
    try:
        return float(re.findall(r"\d+\.\d+|\d+", texto)[0])
    except:
        return 0

def buscar_multiple(patrones, texto):
    for patron in patrones:
        match = re.search(patron, texto, re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(match.lastindex or 0).strip()
    return ""

def limpiar_cliente(nombre):
    if not nombre:
        return ""
    nombre = re.split(r"RUC|Identificaci[oó]n|Direcci[oó]n|Tel[eé]fono", nombre)[0]
    return nombre.strip()

def limpiar_fecha(texto):
    if not texto:
        return ""

    match = re.search(r"\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2}", texto)

    if match:
        fecha = match.group(0)
        try:
            if "/" in fecha:
                return datetime.strptime(fecha, "%d/%m/%Y").strftime("%d/%m/%Y")
            else:
                return datetime.strptime(fecha, "%Y-%m-%d").strftime("%d/%m/%Y")
        except:
            return fecha
    return ""

def extraer_datos(pdf):
    texto = ""

    with pdfplumber.open(pdf) as pdf_file:
        for page in pdf_file.pages:
            t = page.extract_text()
            if t:
                texto += t + "\n"

    cliente_raw = buscar_multiple([
        r"Raz[oó]n Social\s*/\s*Nombres y Apellidos\s*:\s*(.+)",
        r"Raz[oó]n Social\s*:\s*(.+)"
    ], texto)

    cliente = limpiar_cliente(cliente_raw)

    ruc = buscar_multiple([
        r"R\.?U\.?C\.?\s*:\s*(\d{10,13})",
        r"Identificaci[oó]n\s*:\s*(\d{10,13})"
    ], texto)

    if not ruc:
        posibles = re.findall(r"\b\d{13}\b", texto)
        if posibles:
            ruc = posibles[0]

    autorizacion = buscar_multiple([
        r"N[ÚU]MERO\s+DE\s+AUTORIZACI[ÓO]N\s*:\s*(\d+)",
        r"Clave\s+de\s+Acceso\s*:\s*(\d{20,})"
    ], texto)

    if not autorizacion:
        posibles = re.findall(r"\b\d{20,49}\b", texto)
        if posibles:
            autorizacion = posibles[0]

    fecha_raw = buscar_multiple([
        r"Fecha\s+de\s+Emisi[oó]n\s*:\s*([0-9/\-]+)",
        r"FECHA\s+Y\s+HORA\s+DE\s+AUTORIZACI[ÓO]N\s*:\s*([0-9/\-]+)",
        r"Fecha\s*:\s*([0-9/\-]+)"
    ], texto)

    if not fecha_raw:
        posibles = re.findall(r"\d{2}/\d{2}/\d{4}", texto)
        if posibles:
            fecha_raw = posibles[0]

    fecha = limpiar_fecha(fecha_raw)

    factura = buscar_multiple([
        r"Factura\s*No\.?\s*:\s*([\d\-]+)",
        r"(\d{3}-\d{3}-\d{9})"
    ], texto)

    base_0 = limpiar_numero(buscar_multiple([r"0%\s*\$?\s*([\d\.,]+)"], texto))
    base_15 = limpiar_numero(buscar_multiple([r"(?:12%|15%)\s*\$?\s*([\d\.,]+)"], texto))
    propina = limpiar_numero(buscar_multiple([r"PROPINA\s*\$?\s*([\d\.,]+)"], texto))

    iva = round(base_15 * 0.15, 2) if base_15 > 0 else 0

    return {
        "FECHA": fecha if fecha else "NO DETECTADO",
        "CLIENTE": cliente if cliente else "NO DETECTADO",
        "RUC": ruc if ruc else "NO DETECTADO",
        "FACT": factura if factura else "NO DETECTADO",
        "AUTORIZACION": autorizacion if autorizacion else "NO DETECTADO",
        "NO OBJETO": "",
        "EXCENTO IVA": "",
        "BASE 0%": base_0,
        "BASE 15%": base_15,
        "PROPINA": propina,
        "IVA": iva,
        "TOTAL": "",
        "N° RETENCION": "NA",
        "10% R.FTE": "",
        "100% R. IVA": "",
        "TOTAL RETENCION": "",
        "POR COBRAR": ""
    }

if uploaded_files:
    data = []

    for file in uploaded_files:
        try:
            data.append(extraer_datos(file))
        except Exception as e:
            st.error(f"Error en {file.name}: {e}")

    df = pd.DataFrame(data)
    st.dataframe(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "VENTAS FEBRERO"

    headers = list(df.columns)
    ws["A1"] = "VENTAS FEBRERO"
    ws.append(headers)

    amarillo = PatternFill(start_color="FFFF00", fill_type="solid")
    verde = PatternFill(start_color="92D050", fill_type="solid")

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col_num, col_name in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.fill = amarillo
        cell.font = Font(bold=True)
        cell.border = borde

        if col_name == "POR COBRAR":
            cell.fill = verde

    for i, row in enumerate(df.itertuples(index=False), start=3):
        ws.append(row)

        b0 = chr(64 + headers.index("BASE 0%") + 1)
        b15 = chr(64 + headers.index("BASE 15%") + 1)
        prop = chr(64 + headers.index("PROPINA") + 1)
        iva_l = chr(64 + headers.index("IVA") + 1)
        tot = chr(64 + headers.index("TOTAL") + 1)
        pc = chr(64 + headers.index("POR COBRAR") + 1)

        ws[f"{tot}{i}"] = f"={b0}{i}+{b15}{i}+{prop}{i}+{iva_l}{i}"
        ws[f"{pc}{i}"] = f"={tot}{i}"

        for col in range(1, len(headers) + 1):
            ws.cell(row=i, column=col).border = borde

    # -------- FILA DE TOTALES --------
    total_row = len(df) + 3

    ws.cell(row=total_row, column=1, value="TOTAL")

    columnas_sumar = ["BASE 0%", "BASE 15%", "PROPINA", "IVA", "TOTAL", "POR COBRAR"]

    for col_name in columnas_sumar:
        col_letter = chr(64 + headers.index(col_name) + 1)
        ws[f"{col_letter}{total_row}"] = f"=SUM({col_letter}3:{col_letter}{total_row-1})"

    # PINTAR FILA TOTAL
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = amarillo
        cell.font = Font(bold=True)
        cell.border = borde

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.download_button(
        "📥 Descargar Excel FINAL",
        data=output,
        file_name="ventas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
